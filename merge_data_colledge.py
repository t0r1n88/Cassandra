"""
Скрипт для обработки данных по колледжам
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import os

# filename = 'data/Приложение_№_1_Чеченская_Республика_01_12.xlsx'
# filename = 'data/Ингушетия Приложение_№_1.xlsx'
dir_name = 'data/mon'

# Создаем итоговый файл excel
wb = openpyxl.Workbook()
# создаем листы
ren_sheet = wb['Sheet']
ren_sheet.title = 'Свод-студенты'
wb.create_sheet(title='Свод-кадры', index=1)
wb.create_sheet(title='Свод-финансы и МТБ', index=2)

# Создаем итоговые датафреймы
students_df = pd.DataFrame(columns=range(22))
kadr_df = pd.DataFrame(columns=range(18))
fin_df = pd.DataFrame(columns=range(14))

# получаем заголовок первой таблицы, чтобы не пришлось потом руками добавлять заголовок
for dirpath, dirnames, filenames in os.walk(dir_name):
    for filename in filenames:
        if filename.endswith('.xlsx'):
            # Получаем название файла без расширения
            name_file = filename.split('.xlsx')[0]
            print(name_file)
            temb_wb = load_workbook(filename=f'{dirpath}/{filename}', read_only=True)  # загружаем файл
            students_name_sheet = 'Не найден лист с таким названием'
            kadr_name_sheet = 'Не найден лист с таким названием'
            fin_name_sheet = 'Не найден лист с таким названием'

            for name_sheet in temb_wb.sheetnames:  # получаем названия листов, поскольку они могут быть с большой или маленькой буквы
                if 'денты' in name_sheet:
                    students_name_sheet = name_sheet
                if 'адры' in name_sheet:
                    kadr_name_sheet = name_sheet
                if 'МТБ' in name_sheet:
                    fin_name_sheet = name_sheet
                else:
                    continue
            # Создаем датафреймы
            temp_stud_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=students_name_sheet)

            temp_kadr_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=kadr_name_sheet)

            temp_fin_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=fin_name_sheet)
            # Добавляем колонку с названием файла откуда взяты данные
            # temp_stud_df.insert(0, 'Откуда взяты данные', name_file)
            # temp_kadr_df.insert(0, 'Откуда взяты данные', name_file)
            # temp_fin_df.insert(0, 'Откуда взяты данные', name_file)
            temp_stud_df['Откуда взяты данные'] = name_file
            temp_kadr_df['Откуда взяты данные'] = name_file
            temp_fin_df['Откуда взяты данные'] = name_file

            columns_stud_df = temp_stud_df.columns  # получаем заголовки
            columns_kadr_df = temp_kadr_df.columns  # получаем заголовки
            columns_fin_df = temp_fin_df.columns  # получаем заголовки
            break

for dirpath, dirnames, filenames in os.walk(dir_name):
    for filename in filenames:
        if filename.endswith('.xlsx'):
            # Получаем название файла без расширения
            name_file = filename.split('.xlsx')[0]
            print(name_file)
            temb_wb = load_workbook(filename=f'{dirpath}/{filename}', read_only=True)  # загружаем файл
            students_name_sheet = 'Не найден лист с таким названием'
            kadr_name_sheet = 'Не найден лист с таким названием'
            fin_name_sheet = 'Не найден лист с таким названием'

            for name_sheet in temb_wb.sheetnames:  # получаем названия листов, поскольку они могут быть с большой или маленькой буквы
                if 'денты' in name_sheet:
                    students_name_sheet = name_sheet
                if 'адры' in name_sheet:
                    kadr_name_sheet = name_sheet
                if 'МТБ' in name_sheet:
                    fin_name_sheet = name_sheet
                else:
                    continue
            # Создаем датафреймы
            temp_stud_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=students_name_sheet, skiprows=1,
                                         header=None)

            temp_kadr_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=kadr_name_sheet, skiprows=1, header=None)

            temp_fin_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=fin_name_sheet, skiprows=1, header=None)
            print(temp_stud_df.shape)
            print(temp_kadr_df.shape)
            print(temp_fin_df.shape)

            # Добавляем колонку с названием файла откуда взяты данные
            temp_stud_df['Откуда взяты данные'] = name_file
            temp_kadr_df['Откуда взяты данные'] = name_file
            temp_fin_df['Откуда взяты данные'] = name_file

            # сортируем, чтобы было удобнее смотреть

            # temp_stud_df.sort_values(by=[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22],
            #                          ascending=False, inplace=True)
            temp_stud_df.sort_values(by=[1, 2, 3],
                                     ascending=False, inplace=True)
            temp_kadr_df.sort_values(by=list(temp_kadr_df.columns), ascending=False, inplace=True)
            temp_fin_df.sort_values(by=list(temp_fin_df.columns), ascending=False, inplace=True)

            students_df = pd.concat([students_df, temp_stud_df], ignore_index=True)
            kadr_df = pd.concat([kadr_df, temp_kadr_df], ignore_index=True)
            fin_df = pd.concat([fin_df, temp_fin_df], ignore_index=True)

students_df.columns = columns_stud_df
kadr_df.columns = columns_kadr_df
fin_df.columns = columns_fin_df

# записываем в соответствующий лист
for r in dataframe_to_rows(students_df, index=False, header=True):
    wb['Свод-студенты'].append(r)
for r in dataframe_to_rows(kadr_df, index=False, header=True):
    wb['Свод-кадры'].append(r)
for r in dataframe_to_rows(fin_df, index=False, header=True):
    wb['Свод-финансы и МТБ'].append(r)

t = time.localtime()
current_time = time.strftime('%H_%M_%S', t)
# Сохраняем итоговый файл
wb.save(f'Свод от {current_time}.xlsx')
