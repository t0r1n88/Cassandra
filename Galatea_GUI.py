"""
Скрипт для обработки данных по колледжам
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import os

import tkinter
import sys
import os
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import time
# pd.options.mode.chained_assignment = None  # default='warn'
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller
    Функция чтобы логотип отображался"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def select_folder_data():
    """
    Функция для выбора папки c данными
    :return:
    """
    global path_folder_data
    path_folder_data = filedialog.askdirectory()

def select_end_folder():
    """
    Функция для выбора конечной папки куда будут складываться итоговые файлы
    :return:
    """
    global path_to_end_folder
    path_to_end_folder = filedialog.askdirectory()

def select_file_docx():
    """
    Функция для выбора файла Word
    :return: Путь к файлу шаблона
    """
    global file_docx
    file_docx = filedialog.askopenfilename(
        filetypes=(('Word files', '*.docx'), ('all files', '*.*')))

def select_file_data_xlsx():
    """
    Функция для выбора файла с данными на основе которых будет генерироваться документ
    :return: Путь к файлу с данными
    """
    global file_data_xlsx
    # Получаем путь к файлу
    file_data_xlsx = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))

def select_files_data_xlsx():
    """
    Функция для выбора нескоьких файлов с данными на основе которых будет генерироваться документ
    :return: Путь к файлу с данными
    """
    global files_data_xlsx
    # Получаем путь файлы
    files_data_xlsx = filedialog.askopenfilenames(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def processing_data():
    """
    Фугкция для обработки данных
    :return:
    """
    # Создаем итоговый файл excel
    wb = openpyxl.Workbook()
    # создаем листы
    ren_sheet = wb['Sheet']
    ren_sheet.title = 'Свод-студенты'
    wb.create_sheet(title='Свод-кадры', index=1)
    wb.create_sheet(title='Свод-финансы и МТБ', index=2)

    # Создаем итоговые датафреймы
    students_df = pd.DataFrame(columns=range(22))
    kadr_df = pd.DataFrame(columns=range(18))
    fin_df = pd.DataFrame(columns=range(14))

    # получаем заголовок первой таблицы, чтобы не пришлось потом руками добавлять заголовок
    for dirpath, dirnames, filenames in os.walk(path_folder_data):
        for filename in filenames:
            if filename.endswith('.xlsx'):
                # Получаем название файла без расширения
                name_file = filename.split('.xlsx')[0]
                temb_wb = load_workbook(filename=f'{dirpath}/{filename}', read_only=True)  # загружаем файл
                students_name_sheet = 'Не найден лист с таким названием'
                kadr_name_sheet = 'Не найден лист с таким названием'
                fin_name_sheet = 'Не найден лист с таким названием'

                for name_sheet in temb_wb.sheetnames:  # получаем названия листов, поскольку они могут быть с большой или маленькой буквы
                    if 'денты' in name_sheet:
                        students_name_sheet = name_sheet
                    if 'адры' in name_sheet:
                        kadr_name_sheet = name_sheet
                    if 'МТБ' in name_sheet:
                        fin_name_sheet = name_sheet
                    else:
                        continue
                # Создаем датафреймы
                temp_stud_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=students_name_sheet)

                temp_kadr_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=kadr_name_sheet)

                temp_fin_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=fin_name_sheet)
                # Добавляем колонку с названием файла откуда взяты данные
                # temp_stud_df.insert(0, 'Откуда взяты данные', name_file)
                # temp_kadr_df.insert(0, 'Откуда взяты данные', name_file)
                # temp_fin_df.insert(0, 'Откуда взяты данные', name_file)
                temp_stud_df['Откуда взяты данные'] = name_file
                temp_kadr_df['Откуда взяты данные'] = name_file
                temp_fin_df['Откуда взяты данные'] = name_file

                columns_stud_df = temp_stud_df.columns  # получаем заголовки
                columns_kadr_df = temp_kadr_df.columns  # получаем заголовки
                columns_fin_df = temp_fin_df.columns  # получаем заголовки
                break

    try:
        for dirpath, dirnames, filenames in os.walk(path_folder_data):
            for filename in filenames:
                if filename.endswith('.xlsx'):
                    # Получаем название файла без расширения
                    name_file = filename.split('.xlsx')[0]
                    print(name_file)
                    temb_wb = load_workbook(filename=f'{dirpath}/{filename}')  # загружаем файл
                    students_name_sheet = 'Не найден лист с таким названием'
                    kadr_name_sheet = 'Не найден лист с таким названием'
                    fin_name_sheet = 'Не найден лист с таким названием'

                    for name_sheet in temb_wb.sheetnames:  # получаем названия листов, поскольку они могут быть с большой или маленькой буквы
                        if 'денты' in name_sheet:
                            students_name_sheet = name_sheet
                        if 'адры' in name_sheet:
                            kadr_name_sheet = name_sheet
                        if 'МТБ' in name_sheet:
                            fin_name_sheet = name_sheet
                        else:
                            continue
                    # Создаем датафреймы
                    temp_stud_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=students_name_sheet, skiprows=1,
                                                 header=None)

                    temp_kadr_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=kadr_name_sheet, skiprows=1,
                                                 header=None)

                    temp_fin_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=fin_name_sheet, skiprows=1, header=None)
                    print(temp_stud_df.shape)
                    print(temp_kadr_df.shape)
                    print(temp_fin_df.shape)

                    # Добавляем колонку с названием файла откуда взяты данные
                    temp_stud_df['Откуда взяты данные'] = name_file
                    temp_kadr_df['Откуда взяты данные'] = name_file
                    temp_fin_df['Откуда взяты данные'] = name_file


                    students_df = pd.concat([students_df, temp_stud_df], ignore_index=True)
                    kadr_df = pd.concat([kadr_df, temp_kadr_df], ignore_index=True)
                    fin_df = pd.concat([fin_df, temp_fin_df], ignore_index=True)

        students_df.columns = columns_stud_df
        kadr_df.columns = columns_kadr_df
        fin_df.columns = columns_fin_df

        # записываем в соответствующий лист
        for r in dataframe_to_rows(students_df, index=False, header=True):
            wb['Свод-студенты'].append(r)
        for r in dataframe_to_rows(kadr_df, index=False, header=True):
            wb['Свод-кадры'].append(r)
        for r in dataframe_to_rows(fin_df, index=False, header=True):
            wb['Свод-финансы и МТБ'].append(r)

        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        # Сохраняем итоговый файл
        wb.save(f'{path_to_end_folder}/Свод от {current_time}.xlsx')
    except PermissionError:
        messagebox.showerror('ЦОПП Бурятия','Закройте файлы Excel с данными!!! Ну то есть те которые вы хотите обработать')
    else:
        messagebox.showinfo('ЦОПП Бурятия','Обработка завершена!')


if __name__ == '__main__':
    window = Tk()
    window.title('ЦОПП Бурятия')
    window.geometry('700x860')
    window.resizable(False, False)


    # Создаем объект вкладок

    tab_control = ttk.Notebook(window)

    # Создаем вкладку обработки данных для Приложения 6
    tab_report_6 = ttk.Frame(tab_control)
    tab_control.add(tab_report_6, text='Скрипт №1')
    tab_control.pack(expand=1, fill='both')
    # Добавляем виджеты на вкладку Создание образовательных программ
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_report_6,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img = resource_path('logo.png')

    img = PhotoImage(file=path_to_img)
    Label(tab_report_6,
          image=img
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем кнопку Выбрать файл с данными
    btn_choose_data = Button(tab_report_6, text='1) Выберите папку с данными', font=('Arial Bold', 20),
                          command=select_folder_data
                          )
    btn_choose_data.grid(column=0, row=2, padx=10, pady=10)

    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder = Button(tab_report_6, text='2) Выберите конечную папку', font=('Arial Bold', 20),
                                       command=select_end_folder
                                       )
    btn_choose_end_folder.grid(column=0, row=3, padx=10, pady=10)

    #Создаем кнопку обработки данных

    btn_proccessing_data = Button(tab_report_6, text='3) Обработать данные', font=('Arial Bold', 20),
                                       command=processing_data
                                       )
    btn_proccessing_data.grid(column=0, row=4, padx=10, pady=10)

    window.mainloop()