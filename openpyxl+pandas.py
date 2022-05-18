import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
wb = load_workbook('test.xlsx')  # load as openpyxl workbook; useful to keep the original layout
                                 # which is discarded in the following dataframe
df = pd.read_excel('test.xlsx')  # load as dataframe (modifications will be easier with pandas API!)
ws = wb.active
df.iloc[1, 1] = 'hello world'    # modify a few things
rows = dataframe_to_rows(df, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)
wb.save('test2.xlsx')

